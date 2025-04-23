import os
from ..core.interfaces import DocumentServiceInterface
from ..core.models.document import Document
from ..utils.file_utils import file_exists, compute_file_hash, clear_directory
from ..utils.text_processing import preprocess_text
from ..core.dtos import DocumentDetail
from langchain_community.document_loaders import PDFPlumberLoader, CSVLoader, Docx2txtLoader
from openpyxl import load_workbook
import logging
from datetime import datetime
from pymongo import MongoClient
from ..config.config import Config  # Assuming you have a Config class
from ..core.models.document_detail import DocumentDetailModel
from ..services.uploadthing_service import UploadthingService

class DocumentService(DocumentServiceInterface):
    def __init__(self, document_dirs):
        self.document_dirs = document_dirs
        for file_type, directory in self.document_dirs.items():
            os.makedirs(directory, exist_ok=True)
        self.document_detail_model = DocumentDetailModel() # Instantiate the MongoDB model
        self.uploadthing_service = UploadthingService() # Instantiate Uploadthing service

    def list_documents(self, file_type: str) -> list[str]:
        directory = self.document_dirs.get(file_type)
        if directory:
            return [f for f in os.listdir(directory) if f.endswith(self._get_extension(file_type))]
        return []

    def list_document_details(self) -> list[dict]:
        # all_details = []
        # for file_type, directory in self.document_dirs.items():
        #     extension = self._get_extension(file_type)
        #     if directory:
        #         for filename in os.listdir(directory):
        #             if filename.endswith(extension):
        #                 file_path = os.path.join(directory, filename)
        #                 try:
        #                     size = os.path.getsize(file_path)
        #                     # Simulate upload date (can be improved to fetch actual metadata if needed)
        #                     mtime = os.path.getmtime(file_path)
        #                     upload_date = datetime.fromtimestamp(mtime).isoformat()
        #                     detail = DocumentDetail(
        #                         filename=filename,
        #                         size=size,
        #                         type=file_type,
        #                         upload_date=upload_date
        #                     )
        #                     all_details.append(detail.__dict__)
        #                 except Exception as e:
        #                     logging.error(f"Error getting details for {filename}: {e}")
        # return all_details
        return self.document_detail_model.list_document_details()

    def _get_extension(self, file_type: str) -> str:
        if file_type == "pdf":
            return ".pdf"
        elif file_type == "docx":
            return ".docx"
        elif file_type == "csv":
            return ".csv"
        elif file_type == "xlsx":
            return ".xlsx"
        return ""

    def get_document_dir(self, file_type: str) -> str | None:
        return self.document_dirs.get(file_type)

    def upload_document(self, file) -> dict:
        file_name = file.filename
        file_extension = os.path.splitext(file_name)[1].lower()
        file_type = ""

        if file_extension == ".pdf":
            file_type = "pdf"
        elif file_extension == ".docx":
            file_type = "docx"
        elif file_extension == ".csv":
            file_type = "csv"
        elif file_extension == ".xlsx":
            file_type = "xlsx"
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")

        save_dir = self.document_dirs[file_type]
        save_file = os.path.join(save_dir, file_name)

        if file_exists(save_file):
            print ("File already exists.")
            raise ValueError("File already exists.")

        try:
            file_hash = compute_file_hash(file)
        except Exception as e:
            raise Exception(f"Error computing file hash: {e}") from e

        existing_files = self.list_documents(file_type)
        for existing_file in existing_files:
            existing_file_path = os.path.join(save_dir, existing_file)
            try:
                if compute_file_hash(open(existing_file_path, 'rb')) == file_hash:
                    raise ValueError("File with identical content already exists.")
            except Exception as e:
                raise Exception(f"Error checking existing files: {e}") from e

        try:
            file.save(save_file)
        except Exception as e:
            raise Exception(f"Error saving file: {e}") from e

        docs = []
        is_structured = True

        try:
            if file_type == "pdf":
                loader = PDFPlumberLoader(save_file)
                loaded_docs = loader.load_and_split()
                docs = [Document(page_content=preprocess_text(doc.page_content), metadata={"source": file_name}) for doc in loaded_docs]
            elif file_type == "docx":
                text = Docx2txtLoader(save_file).load()[0].page_content
                docs = [Document(page_content=preprocess_text(text), metadata={"source": file_name})]
            elif file_type == "csv":
                loader = CSVLoader(save_file, encoding="utf-8")
                loaded_docs = loader.load()
                docs = [Document(page_content=preprocess_text(doc.page_content), metadata={"source": file_name}) for doc in loaded_docs]
            elif file_type == "xlsx":
                workbook = load_workbook(save_file)
                text = ""
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        row_values = [cell.value for cell in row if cell.value is not None]
                        text += " ".join(map(str, row_values)) + "\n"
                docs = [Document(page_content=preprocess_text(text), metadata={"source": file_name})]
            else:
                is_structured = False
        except Exception as e:
            logging.warning(f"Error loading content from {file_type} file: {e}")
            is_structured = False
            docs = [Document(page_content="Error loading document content.", metadata={"source": file_name})]

        # --- Integration with Uploadthing and MongoDB ---
        uploadthing = None
        try:
            uploadthing = self.uploadthing_service.upload(file, os.path.getsize(save_file))
            key_value = uploadthing['key']
            # print("uploadthing", uploadthing)
            logging.info(type(key_value))
            # Save document details to MongoDB
            upload_date = datetime.now()
            self.document_detail_model.insert_document_detail(
                name=file_name,
                size=os.path.getsize(save_file),
                upload_date=upload_date,
                file_type=file_type,
                url=uploadthing['url'],
                key=uploadthing['key']
            )
            logging.info(f"Document details saved to MongoDB for: {file_name} with URL: {uploadthing}")

        except Exception as e:
            logging.error(f"Error during Uploadthing upload or saving to MongoDB: {e}")

        return {
            "filename": file_name,
            "doc_len": len(docs),
            "chunks": docs,
            "is_structured": is_structured,
            "file_type": file_type,
        }

    def delete_document(self, file_id: str, file_name: str, file_type: str) -> None:
        directory = self.document_dirs.get(file_type)
        self.document_detail_model.delete_document_detail_by_id(file_id)
        if directory:
            file_path = os.path.join(directory, file_name)
            if os.path.exists(file_path):
                os.remove(file_path)
                logging.info(f"Successfully deleted file: {file_path}")
            else:
                logging.warning(f"File not found: {file_path}")
                raise FileNotFoundError(f"File not found: {file_path}")
        else:
            logging.warning(f"Invalid file type: {file_type}")
            raise ValueError(f"Invalid file type: {file_type}")

    def clear_document_directory(self, file_type: str) -> None:
        directory = self.document_dirs.get(file_type)
        if directory:
            clear_directory(directory)
            self.document_detail_model.clear_db()
        else:
            logging.warning(f"Invalid file type: {file_type}")
            raise ValueError(f"Invalid file type: {file_type}")

    # --- Hypothetical function for Uploadthing integration ---
    def upload_to_uploadthing(self, file):
        """
        This is a placeholder for your actual Uploadthing integration logic.
        It should handle the upload and return the URL.
        """
        # Replace this with your actual implementation
        return f"https://uploadthing.com/your-upload/{file.filename}"